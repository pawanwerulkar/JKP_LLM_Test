import json
from copyreg import pickle
from operator import index
import re
import openai
import ast
from cssselect.parser import Class
from openpyxl import load_workbook, Workbook
import config
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# from pinecone_utils import query_pinecone2

openai.api_key = config.OPEN_API_KEY
# Initialize OpenAI API key
 # Replace with your actual API key

import openai


def get_llm_response(question, response_text):
    try:
        openai.api_key = config.OPEN_API_KEY
        # Construct the prompt for LLM
        prompt = f"Grade the following response from 0-10 on how relevent it is to the following search request.Only respond with a number from 0 to 10: {question} response: {response_text}"

        # API call using the chat models (gpt-3.5-turbo or gpt-4)
        response = openai.chat.completions.create(
            model="gpt-4o",  # You can use 'gpt-4' if available
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ]
        )

        print(f"question:{question}")
        print(f"response_text: {response_text}")
        # Extracting the response from LLM
        llm_response = response.choices[0].message.content
        print(llm_response)
        return llm_response


    except openai.OpenAIError as e:
        # Handling OpenAI API specific errors
        print(f"Error with OpenAI API: {e}")
        return None
    except Exception as e:
        # General error handling
        print(f"Error getting LLM response: {e}")
        return None




def calculate_video_matching_and_snippet_overlap(expected_start_time, actual_start_time, expected_end_time,
                                                 actual_end_time):
    """
    Compare the expected and actual start/end times and return whether they match.
    """
    if not isinstance(expected_start_time, str) or not isinstance(actual_start_time, str):
        print("Both expected_start_time and actual_start_time must be strings.")
        return False, False  # Invalid case

    if not isinstance(expected_end_time, str) or not isinstance(actual_end_time, str):
        print("Both expected_end_time and actual_end_time must be strings.")
        return False, False  # Invalid case

    # Case insensitive match check for start and end times
    start_time_match = expected_start_time.lower() == actual_start_time.lower()
    end_time_match = expected_end_time.lower() == actual_end_time.lower()

    return start_time_match, end_time_match

# Function to calculate similarity score using cosine similarity
def calculate_similarity_score(response_text, ground_truth_text):
    try:
        # Combine the two texts (response and ground truth) for vectorization
        documents = [response_text, ground_truth_text]

        # Create a TfidfVectorizer for text vectorization
        vectorizer = TfidfVectorizer()

        # Transform the documents into TF-IDF vectors
        tfidf_matrix = vectorizer.fit_transform(documents)

        # Calculate cosine similarity between the two vectors
        similarity_matrix = cosine_similarity(tfidf_matrix[0:1], tfidf_matrix[1:2])

        # Return the similarity score
        return similarity_matrix[0][0]  # Cosine similarity value between 0 and 1

    except Exception as e:
        print(f"Error calculating similarity score: {e}")
        return 0  # Return 0 if there's an error



def append_summary_to_new_workbook(input_file_path, output_file_path, results_sheet_name="Test Results", overall_sheet_name="Test Report Structure (Overall)"):
    try:
        # Load the input workbook
        input_workbook = load_workbook(input_file_path)

        # Check if the required sheets exist in the workbook
        if results_sheet_name not in input_workbook.sheetnames:
            raise ValueError(f"Sheet '{results_sheet_name}' not found in the Excel file.")
        if overall_sheet_name not in input_workbook.sheetnames:
            raise ValueError(f"Sheet '{overall_sheet_name}' not found in the Excel file.")

        # Get the sheets from the input workbook
        results_sheet = input_workbook[results_sheet_name]
        overall_sheet = input_workbook[overall_sheet_name]

        # Initialize a dictionary to group rows by index name
        index_data = {}

        # Read the headers (assume the first row is the header) for the Test Results sheet
        headers = [cell.value for cell in results_sheet[1]]

        # Find the necessary columns in the Test Results sheet
        index_name_col = headers.index("Index Name")
        result_col = headers.index("Result")  # Assuming 'Result' column exists
        category_col = headers.index("Category")  # Assuming 'Category' column exists
        expected_start_time_col = headers.index("Ground Truth Snippet Start")
        actual_start_time_col = headers.index("Response Snippet Start")
        expected_end_time_col = headers.index("Ground Truth Snippet End")
        actual_end_time_col = headers.index("Response Snippet End")
        question_col = headers.index("Question")
        response_text_col = headers.index("Response Text")  # Assuming 'Response Text' column exists
        ground_truth_video_title_col = headers.index("Ground Truth Video Title")
        top_response_video_title_col = headers.index("Response Video Title")
        ground_truth_text_col = headers.index("Ground Truth Text")
        similarity_score_col = headers.index("Similarity Score")
        llm_response_col = headers.index("LLM Response")
        response_video_link_col = headers.index("Response Video")
        detail_respose_col = headers.index("Detail Response")


        # Iterate over the rows (starting from the second row to exclude the header)
        for row in results_sheet.iter_rows(min_row=2, max_row=results_sheet.max_row, min_col=1, max_col=results_sheet.max_column):
            index_name = row[index_name_col].value
            if index_name:  # Only process rows with an index name
                if index_name not in index_data:
                    index_data[index_name] = []
                index_data[index_name].append([cell.value for cell in row])  # Store row data

        # Create a new workbook for the final report
        final_workbook = Workbook()



        # For each index, create a new sheet and append data
        for index_name, rows in index_data.items():
            # Create a new sheet for each index name
            final_sheet = final_workbook.create_sheet(title=index_name)
            final_sheet.append([])



            #Initialized the pass and fail count to track pass and fail count
            pass_count = 0
            fail_count= 0

            #Iterate through the rows and check the "Result" column
            for row in rows:
                result = row[result_col]
                if result == "Pass":
                    pass_count += 1
                elif result == "Fail":
                    fail_count+=1

            #calculate the overall percentage
            total_test = pass_count+fail_count
            if total_test > 0:
                overall_percentage = (pass_count/total_test)*100
            else:
                overall_percentage = 0

            # headers for table 1
            final_sheet.append(["Index Name","Overall"])
            # Append the row with index name and overall percentage

            overall_row = [index_name, f"{overall_percentage:.2f}%"]  # Format percentage to 2 decimal places
            final_sheet.append(overall_row)


            final_sheet.append([])  # Add an empty row

            # Add the Category Table below the Overall Index Performance Table (Table 2)
            final_sheet.append(["By Category"])
            final_sheet.append(["Category", "Pass Rate (%)"])  # Table 2 Header

            # Group rows by Category
            category_data = {}
            for row in rows:
                category = row[category_col]  # Get the category from the row
                result = row[result_col]  # Get the result from the row

                if category not in category_data:
                    category_data[category] = {"total": 0, "pass": 0}

                category_data[category]["total"] += 1
                if result.lower() == "pass":  # Check if the result is 'Pass'
                    category_data[category]["pass"] += 1

            # Append Category pass rate data
            for category, counts in category_data.items():
                pass_rate = (counts["pass"] / counts["total"]) * 100 if counts["total"] > 0 else 0
                final_sheet.append([category, f"{pass_rate:.2f}%"])

            # Add an empty row after the Category Table
            final_sheet.append([])  # Add an empty row

            # Add the "By Question" Table (Table 3)
            final_sheet.append(["By Question"])
            final_sheet.append(["Question","Result Output", "Result", "Video Matching", "Snippet TS Overlap", "LLM Response","Smilarity Score"])  # Table 3 Header

            for row in rows:
                question = row[question_col]  # Question data
                response_text = row[response_text_col]  # Response Text
                result = row[result_col]
                question = row[question_col]  # Question data
                expected_start_time = row[expected_start_time_col]  # Start Time data
                actual_start_time = row[actual_start_time_col]  # Actual Start Time
                expected_end_time = row[expected_end_time_col]  # End Time data
                actual_end_time = row[actual_end_time_col]  # Actual End Time
                similarity_score = row[similarity_score_col]
                llm_response = row[llm_response_col]


                # Get LLM Response (Grading)
                # llm_response = get_llm_response(question, response_text)

                # Calculate Video Matching and Snippet TS Overlap (Assuming your existing function is already there)
                video_matching, snippet_ts_overlap = calculate_video_matching_and_snippet_overlap(
                    row[expected_start_time_col], row[actual_start_time_col], row[expected_end_time_col], row[actual_end_time_col]
                )

                # Based on matching, determine the result
                # result = "Pass" if video_matching and snippet_ts_overlap else "Fail"
                if int(llm_response) >= 7:
                    result = "Pass"
                else:
                    result = "Fail"

                # Append question data with calculated result and LLM response
                final_sheet.append([question,response_text,result, "Yes" if video_matching else "No", "Yes" if snippet_ts_overlap else "No", llm_response,similarity_score])

            # Add an empty row after the "By Question" Table
            final_sheet.append([])  # Add an empty row

            # Add the Deep Dive table header[table 4]
            final_sheet.append(["Deep Dive"])
            final_sheet.append([
                "Category",
                "Question",
                "Result",
                "LLM Response Score",
                "Top Response Similarity Score",
                "Response Text",
                "Top Response Video Title",
                "Top Response Snippet Start to End TS",
                "Response Video Link"
            ])

            # Now append each row of data to the Deep Dive table
            for row in rows:
                category = row[category_col]
                question = row[question_col]
                result = row[result_col]
                llm_response = row[llm_response_col]
                similarity_score = row[similarity_score_col]
                response_text = row[response_text_col]
                top_response_video_title = row[top_response_video_title_col]
                act_start_time = row[actual_start_time_col]
                act_end_time = row[actual_end_time_col]
                actual_start_to_end_TS = f"{act_start_time} to {act_end_time}"
                response_video_link = row[response_video_link_col]
                detail_response = row[detail_respose_col]

                final_sheet.append([
                    category, question, result, llm_response,similarity_score,response_text,top_response_video_title,
                    actual_start_to_end_TS,response_video_link
                    # ground_truth_start_to_end_TS,
                    # top_response_video_title,
                    # actual_start_to_end_TS,  # Replace with actual timestamp data
                    # response_text, snippet_ts_overlap, ground_truth_video_link,
                    # response_video_link, llm_response if llm_response else "N/A"
                ])
                # # print(response_video_link)
                # print(detail_response)
                # print(type(detail_response))
                # # print(json.loads(detail_response))
                #
                pattern = re.compile(
                    r"media_source':\s*'(.*?)'.*?end':\s*'(.*?)'.*?name':\s*'(.*?)'.*?source':\s*'(.*?)'.*?start':\s*'(.*?)'.*?page_content='(.*?)'.*?,\s*(0\.\d+)\)",
                    re.DOTALL,
                )

                # Find all matches
                matches = pattern.findall(detail_response)

                print(f"lenth of matches {matches}")

                # Print the extracted data
                for match in matches:
                    media_source, end, name, source, start, page_content, smilarity_score = match
                    print(f"Media Source: {media_source}")
                    print(f"Start: {start}")
                    print(f"End: {end}")
                    print(f"Name: {name}")
                    print(f"Source: {source}")

                    start_end_ts = (f"start_end_ts {start} to {end}")
                    print(start_end_ts)
                    llm_rsp = get_llm_response(question,page_content)
                    llm_rsp = int(llm_rsp)
                    print(f"llm_resp{llm_rsp}")
                    page_content = page_content[:100]
                    resp_result = ''
                    if llm_rsp >=7:
                        resp_result = "Pass"
                    else:
                        resp_result = "Fail"

                    print(resp_result)

                    print(f"Page Content: {page_content[:100]}...")  # Truncate page content for readability
                    print(f"similarity_score:{similarity_score}")
                    print("=" * 50)
                    final_sheet.append(["N/A","N/A",resp_result,llm_rsp,similarity_score,page_content,name,start_end_ts,source])





                # Placeholder for video matching, ground truth video link, and response video link
                # video_matching = "N/A"  # Implement actual logic
                # snippet_ts_overlap = "N/A"  # Implement actual logic
                # ground_truth_video_link = "N/A"  # Implement actual logic
                # # response_video_link = "N/A"  # Implement actual logic
                # ground_truth_similariry_score = "NA"

                # Video Matching and Snippet TS Overlap function
                video_matching, snippet_ts_overlap = calculate_video_matching_and_snippet_overlap(
                    row[expected_start_time_col], row[actual_start_time_col], row[expected_end_time_col],
                    row[actual_end_time_col]
                )

                # Determine the results for Video Matching and Snippet TS Overlap
                video_matching = "Pass" if video_matching else "Fail"
                snippet_ts_overlap = "Pass" if snippet_ts_overlap else "Fail"
                # Add row data to Deep Dive table





            final_sheet.append([])  # Add an empty row

            # Remove the default sheet created in the new workbook (if any)
        if "Sheet" in final_workbook.sheetnames:
            del final_workbook["Sheet"]

        # Save the final workbook to the specified output file path
        final_workbook.save(output_file_path)
        print(f"Data has been successfully appended to {output_file_path} with separate sheets for each index.")

    except Exception as e:
        print(f"An error occurred: {e}")


# Example usage:
input_file_path = r"C:\Users\test_complete\PycharmProjects\KN_LLM\report\test_report.xlsx"
output_file_path = r"C:\Users\test_complete\PycharmProjects\KN_LLM\report\final_report.xlsx"

append_summary_to_new_workbook(input_file_path, output_file_path)
