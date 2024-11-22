import os
from os.path import basename

from openpyxl import Workbook, load_workbook

from datetime import datetime

current_date = datetime.now().strftime('%y-%m-%d')
# basename,extension = os.pat.splittext()

# Define the path for the report
report_path = os.path.join(os.path.dirname(__file__), "report", fr'C:\Users\test_complete\PycharmProjects\KN_LLM\report\test_report_{current_date}.xlsx')

# # Ensure the report directory exists
# def test_remove_file():
#     report_dir = os.path.dirname(report_path)
#     os.remove(report_path)
#     print("excel removed!")
#     os.remove("report/test_report.html")
#     print("html removed")





_workbook_instance = None  # Global variable for workbook instance


def get_or_create_report():
    global _workbook_instance
    report_dir = "report"
    report_path = os.path.join(report_dir, f"test_report_{current_date}.xlsx")

    # Ensure the report directory exists
    if not os.path.exists(report_dir):
        os.makedirs(report_dir)

    if _workbook_instance is None:
        if os.path.exists(report_path):
            _workbook_instance = load_workbook(report_path)
        else:
            _workbook_instance = Workbook()
            # Create Test Results sheet with headers
            ws_results = _workbook_instance.create_sheet("Test Results", 0)
            ws_results.append([
                "Question","Index Name", "Response Video Title", "Response Text",
                "Response Video", "Response Snippet Start",
                "Response Snippet End", "Ground Truth Video Title",
                "Ground Truth Text", "Ground Truth Video",
                "Ground Truth Snippet Start", "Ground Truth Snippet End",
                "Result"
            ])

            # Create Metrics sheet with headers
            ws_metrics = _workbook_instance.create_sheet("Metrics", 1)
            ws_metrics.append(["Model", "Metric Type", "Accuracy"])

            # Remove default sheet created by Workbook()
            if "Sheet" in _workbook_instance.sheetnames:
                del _workbook_instance["Sheet"]

            _workbook_instance.save(report_path)

    return _workbook_instance


# def calculate_and_save_summary():
#     wb = load_workbook(report_path)
#     ws_results = wb["Test Results"]
#     ws_summary = wb["Summary"]
#
#     total_tests = passed_tests = failed_tests = 0
#     for row in ws_results.iter_rows(min_row=2, values_only=True):
#         total_tests += 1
#         if row[-1].lower() == "pass":
#             passed_tests += 1
#         else:
#             failed_tests += 1
#
#     pass_percentage = (passed_tests / total_tests) * 100 if total_tests else 0
#     ws_summary.append([total_tests, passed_tests, failed_tests, f"{pass_percentage:.2f}%"])
#
#     wb.save(report_path)  # Save after updating summary
#     print("Summary saved in 'Summary' sheet.")

# def calculate_and_save_average_accuracy():
#     wb = load_workbook(report_path)
#     ws_metrics = wb["Metrics"]
#     ws_avg_accuracy = wb["Average Accuracy"]
#
#     model_accuracies = {}
#     for row in ws_metrics.iter_rows(min_row=2, values_only=True):
#         model, _, accuracy = row
#         model_accuracies.setdefault(model, []).append(accuracy)
#
#     for model, accuracies in model_accuracies.items():
#         avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else 0
#         ws_avg_accuracy.append([model, f"{avg_accuracy:.2f}%"])
#
#     wb.save(report_path)  # Save after updating average accuracy
#     print("Average accuracies saved.")

def calculate_and_save_summary():
    wb = load_workbook(report_path)
    ws_results = wb["Test Results"]

    # Check or create 'Summary' sheet
    if "Summary" not in wb.sheetnames:
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Total Tests", "Passed Tests", "Failed Tests", "Pass Percentage"])
    else:
        ws_summary = wb["Summary"]

    total_tests = passed_tests = failed_tests = 0

    # Iterate over test results to calculate summary
    for row in ws_results.iter_rows(min_row=2, values_only=True):
        total_tests += 1
        if row[-1].lower() == "pass":
            passed_tests += 1
        else:
            failed_tests += 1

    pass_percentage = (passed_tests / total_tests) * 100 if total_tests else 0

    # Overwrite or append the summary row
    last_row = ws_summary.max_row
    if last_row > 1:
        ws_summary.delete_rows(2, last_row - 1)
    ws_summary.append([total_tests, passed_tests, failed_tests, f"{pass_percentage:.2f}%"])

    wb.save(report_path)  # Save after updating summary
    print("Summary saved in 'Summary' sheet.")

# def calculate_and_save_average_accuracy():
#     wb = load_workbook(report_path)
#     ws_metrics = wb["Metrics"]
#
#     # Check or create 'Average Accuracy' sheet
#     if "Average Accuracy" not in wb.sheetnames:
#         ws_avg_accuracy = wb.create_sheet("Average Accuracy")
#         ws_avg_accuracy.append(["Model", "Average Accuracy"])
#     else:
#         ws_avg_accuracy = wb["Average Accuracy"]
#
#     model_accuracies = {}
#     for row in ws_metrics.iter_rows(min_row=2, values_only=True):
#         model, _, accuracy = row
#         model_accuracies.setdefault(model, []).append(accuracy)
#
#     # Overwrite or append average accuracy data
#     last_row = ws_avg_accuracy.max_row
#     if last_row > 1:
#         ws_avg_accuracy.delete_rows(2, last_row - 1)
#
#     for model, accuracies in model_accuracies.items():
#         avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else 0
#         ws_avg_accuracy.append([model, f"{avg_accuracy:.2f}%"])
#
#     wb.save(report_path)  # Save after updating average accuracy
#     print("Average accuracies saved.")

def calculate_accuracy(expected_snippet, actual_snippet):
    expected_words = expected_snippet.split()
    actual_words = actual_snippet.split()
    match_count = sum(1 for word in expected_words if word in actual_words)
    return (match_count / len(expected_words)) * 100 if expected_words else 0
# Example Usage:
# get_or_create_report()  # This will ensure the report and directory are created if not present.
# calculate_and_save_summary()  # This will calculate and append the summary to the report.
# calculate_and_save_average_accuracy()  # This will calculate and append average accuracy to the report.
# from datetime import datetime
def calculate_and_save_average_accuracy():
    # Get today's date and format it as YYYY-MM-DD
    # current_date = datetime.now().strftime("%Y-%m-%d")
    #
    # # Define the report filename with the current date
    # report_path = f"test_report_{current_date}.xlsx"
    wb = load_workbook(report_path)
    ws_metrics = wb["Metrics"]

    # Check or create 'Average Accuracy' sheet
    if "Average Accuracy" not in wb.sheetnames:
        ws_avg_accuracy = wb.create_sheet("Average Accuracy")
        ws_avg_accuracy.append(["Module", "Average Accuracy"])
    else:
        ws_avg_accuracy = wb["Average Accuracy"]

    module_accuracies = {}

    # Group accuracies by module
    for row in ws_metrics.iter_rows(min_row=2, values_only=True):
        module, _, accuracy = row
        if module:
            module_accuracies.setdefault(module, []).append(accuracy)

    # Clear existing data before appending new calculations
    last_row = ws_avg_accuracy.max_row
    if last_row > 1:
        ws_avg_accuracy.delete_rows(2, last_row - 1)

    # Calculate and save average accuracy for each module
    for module, accuracies in module_accuracies.items():
        avg_accuracy = sum(accuracies) / len(accuracies) if accuracies else 0
        ws_avg_accuracy.append([module, f"{avg_accuracy:.2f}%"])

    wb.save(report_path)  # Save the updated workbook
    print("Average accuracies saved for each module.")

# def save_report_date():
#     report_path = r'C:\Users\test_complete\PycharmProjects\KN_LLM\report\test_report.xlsx'
#
#     # if not os._exists(report_path):
#     #     print("specified report is not present on the location.")
#     wb = load_workbook(report_path)
#     current_date = datetime.now().strftime('%y-%m-%d')
#
#     basename, extension = os.path.splitext(report_path)
#
#     new_report_name = f"test_report_{current_date}{extension}"
#     print(f"report saved successfully {new_report_name}")
#     report_path_split = report_path.split("\\")
#     report_path_split[-1] = new_report_name
#     report_path = '\\'.join(report_path_split)
#     wb.save(report_path)
#
#
#     return new_report_name