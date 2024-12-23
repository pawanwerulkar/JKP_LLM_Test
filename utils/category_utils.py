from openpyxl import load_workbook
from collections import defaultdict


def read_report_and_calculate_pass_fail_percentage(report_file_path):
    # Load the Excel workbook and sheet
    wb = load_workbook(report_file_path)
    sheet = wb["Test Results"]  # Sheet name where the results are stored

    # Define the column indexes for necessary columns (we'll assume the headers are in the first row)
    headers = [cell.value for cell in sheet[1]]
    result_index = headers.index("Result")
    index_name_index = headers.index("Index Name")

    # Initialize a dictionary to store results grouped by Index Name
    index_results = defaultdict(list)

    # Iterate through the rows, starting from the second row (skipping the header)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        index_name = row[index_name_index].value
        result = row[result_index].value

        # Append the result to the respective index name
        if index_name and result:
            index_results[index_name].append(result)

    # Initialize a dictionary to store the pass/fail percentage per index
    pass_fail_percentage = {}

    # Calculate the pass/fail percentage for each index
    for index_name, results in index_results.items():
        total_tests = len(results)
        pass_count = results.count("Pass")
        fail_count = results.count("Fail")

        # Calculate pass percentage
        pass_percentage = (pass_count / total_tests) * 100 if total_tests > 0 else 0
        fail_percentage = 100 - pass_percentage

        # Store the result in the dictionary
        pass_fail_percentage[index_name] = {
            "Pass Count": pass_count,
            "Fail Count": fail_count,
            "Pass Percentage": pass_percentage,
            "Fail Percentage": fail_percentage
        }

    # Return the result dictionary for further processing or printing
    return pass_fail_percentage


# Example usage:
report_file_path = input_file_path = r"C:\Users\test_complete\PycharmProjects\KN_LLM/report/test_report.xlsx"
pass_fail_results = read_report_and_calculate_pass_fail_percentage(report_file_path)

# Print the results
for index_name, result in pass_fail_results.items():
    print(f"Index: {index_name}")
    print(f"  Pass Count: {result['Pass Count']}")
    print(f"  Fail Count: {result['Fail Count']}")
    print(f"  Pass Percentage: {result['Pass Percentage']:.2f}%")
    print(f"  Fail Percentage: {result['Fail Percentage']:.2f}%")
